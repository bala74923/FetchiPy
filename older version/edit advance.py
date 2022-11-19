# from bs4 import BeautifulSoup  not used so removed
# just get rank and fetch invalid or not
import time

import requests
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.styles.borders import BORDER_THIN

start_time = time.time()
CONTEST_NAME = "Weekly contest 307"
#contest_name = "weekly-contest-307"  # inpghggut
contest_name = CONTEST_NAME.lower().strip().replace(' ','-')
# score_array = [3,4,5,6]  # input will be updated soon
path = "C:\\Users\\nobel\\PycharmProjects\\pythonProject\\dataset\\LC Student ID.xlsx"  # path where input is fetched

black = '00000000'
red = 'f54242'
green = '42f572'
yellow = 'ebcf34'
violet = '6d51a6'

thin_border = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000')
)

def searchMaxCol(leet_code_id_row,startCol):
    for col_val in range(startCol,sheet_obj.max_column+1):
        curr_val = sheet_obj.cell(row=leet_code_id_row, column=col_val).value
        if curr_val is None:
            return col_val-1
    return sheet_obj.max_column

def searchMaxRow(startRow,leet_code_id_col):
    for row_val in range(startRow,sheet_obj.max_row+1):
        curr_val = sheet_obj.cell(row=row_val, column=leet_code_id_col).value
        if curr_val is None:
            return row_val-1
    return sheet_obj.max_row

def get_dictionary(obj_from_list,solved_progs):
    mydict = {
        #"name": obj_from_list['username'],
        "rank": obj_from_list['rank'],
        "solved":solved_progs
    }
    return mydict


def is_contains_leetcode_ids(string):
    string = string.lower()

    # return true if cell has string leetcode id
    return string.__contains__("user") and string.__contains__("id")


def get_row_col_position_for_leetcode_id():
    for row_ind in range(1, max_row + 1):
        for col_ind in range(1, sheet_obj.max_column + 1):
            curr_val = sheet_obj.cell(row=row_ind, column=col_ind).value
            if curr_val is not None and is_contains_leetcode_ids(curr_val):
                return [row_ind, col_ind]
    return None


def create_columns():
    # we have to create column names for ranking and solved
    # color
    sheet_obj.cell(row=leetcode_id_cell_position[0], column=col_to_fill_rank).fill = PatternFill(start_color=yellow,
                                                                                                 end_color=yellow,
                                                                                                 fill_type="solid")
    sheet_obj.cell(row=leetcode_id_cell_position[0], column=col_to_fill_solved).fill = PatternFill(start_color=yellow,
                                                                                                   end_color=yellow,
                                                                                                   fill_type="solid")
    # name
    sheet_obj.cell(row=leetcode_id_cell_position[0], column=col_to_fill_rank).value = 'Ranking'
    sheet_obj.cell(row=leetcode_id_cell_position[0], column=col_to_fill_solved).value = 'Solved'

    # border
    sheet_obj.cell(row=leetcode_id_cell_position[0], column=col_to_fill_rank).border = thin_border
    sheet_obj.cell(row=leetcode_id_cell_position[0], column=col_to_fill_solved).border = thin_border


def mention_contest_name(row,col):
    print('row = ', row, 'col = ', col)
    sheet_obj.cell(row=row, column=col).border = thin_border
    sheet_obj.cell(row=row, column=col+1).border = thin_border

    #sheet_obj.cell(row=row, column=col).value = contest_name

    sheet_obj.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    print(sheet_obj.merged_cells.ranges)

    cell = sheet_obj.cell(row=row, column=col)
    cell.value = contest_name


    cell.border = thin_border
    #
    cell.alignment = Alignment(horizontal='center', vertical='center')

#global
total_rank_dict = dict()
total_ranks = []
total_submissions = []

user_row = dict()  # maps users with his row, so we can later update it



final_list = []

# code to fetch total_ranks,total_submissions
page = 1
while True:
    try:
        API_URL_FMT = 'https://leetcode.com/contest/api/ranking/{}/?pagination={}&region=global'
        url = API_URL_FMT.format(contest_name, page)
        resp = requests.get(url).json()
        # print(resp)
        objs = resp['total_rank']
        qns = resp['questions']
        submissions = resp['submissions']
        if len(objs) == 0:
            break
        total_ranks.append(objs)
        total_submissions.append(qns)

        for index,obj in enumerate(objs):
            # print(obj['username'], obj['rank'], calculate_solved_programs(obj['score']))
            obj['username'] = obj['username'].lower()
            curr_obj_solved = f'{len(submissions[index])}/4'
            user_details = get_dictionary(obj, curr_obj_solved)
            total_rank_dict[obj['username']] = user_details
            # print(obj)
        print('page = ', page, ' done')
        page = page + 1
    except Exception as e:
        print(e,' so we cannot fetch details')

for user in total_rank_dict.keys():
    print(user,total_rank_dict[user])

wb_obj = openpyxl.load_workbook(path)
#sheet_obj = wb_obj.active


for sheet_name in wb_obj.sheetnames:
    sheet_obj = wb_obj[sheet_name]
    print(sheet_name,sheet_obj)
    max_row = sheet_obj.max_row
    max_col = sheet_obj.max_column

    users = []  # for all username
    attended = []  # users who attended

    # finding all positions to fill (no user given input)
    leetcode_id_cell_position = get_row_col_position_for_leetcode_id()
    leetcode_id_column_name = leetcode_id_cell_position[1]  # 6-starts from 1
    leetcode_id_row_start_position = leetcode_id_cell_position[0] + 1  # 4- starts from 1
    max_col = searchMaxCol(leetcode_id_cell_position[0],leetcode_id_column_name)
    max_row = searchMaxRow(leetcode_id_cell_position[0],leetcode_id_column_name)
    col_to_fill_rank = max_col + 1
    col_to_fill_solved = max_col + 2

    print(leetcode_id_cell_position)
    print(max_col)

    mention_contest_name(leetcode_id_cell_position[0] - 1, max_col + 1)
    create_columns()  # create columns
    for row_val in range(leetcode_id_row_start_position, max_row + 1):
        curr_name = str(sheet_obj.cell(row=row_val, column=leetcode_id_column_name).value)
        curr_name = curr_name.lower().strip()# case not sensitive

        rank_cell = sheet_obj.cell(row=row_val, column=col_to_fill_rank)
        solved_cell = sheet_obj.cell(row=row_val, column=col_to_fill_solved)
        # default red for all
        rank_cell.fill = PatternFill(start_color=red, end_color=red, fill_type="solid")
        solved_cell.fill = PatternFill(start_color=red, end_color=red, fill_type="solid")

        # default NA fill for ALL
        rank_cell.value = 'Not Attended'
        solved_cell.value = 'Not Attended'

        # default black border for ALL
        rank_cell.border = thin_border
        solved_cell.border = thin_border

        # default center alignment for ALL
        rank_cell.alignment = Alignment(horizontal='center', vertical='center')
        solved_cell.alignment = Alignment(horizontal='center', vertical='center')

        # print(row_val, curr_name)
        if total_rank_dict.get(curr_name,None) is not None:
            #curr_name = curr_name.lower().strip()
            #users.append(curr_name)
            print(row_val, curr_name)
            # default red for all
            rank_cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
            solved_cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")

            # default NA fill for ALL
            rank_cell.value = total_rank_dict[curr_name]['rank']
            solved_cell.value = total_rank_dict[curr_name]['solved']
        else:
            url = f'https://leetcode.com/graphql/?query=query{{ userContestRankingHistory(username: "{curr_name}") {{ attended trendDirection problemsSolved totalProblems finishTimeInSeconds rating ranking contest {{ title startTime }} }} }}'
            try:
                resp = requests.get(url).json()['errors'] # if errors is present then that is invalid
                # default red for all
                rank_cell.fill = PatternFill(start_color=violet, end_color=violet, fill_type="solid")
                solved_cell.fill = PatternFill(start_color=violet, end_color=violet, fill_type="solid")

                # default NA fill for ALL
                rank_cell.value = 'Invalid ID'
                solved_cell.value = 'Invalid ID'
            except Exception as e:           #     if errors not present then valid
                pass


    #  print(users)









# #old one
# try:
#     page = 1
#     while len(attended) != len(users):
#         API_URL_FMT = 'https://leetcode.com/contest/api/ranking/{}/?pagination={}&region=global'
#         url = API_URL_FMT.format(contest_name, page)
#         resp = requests.get(url).json()
#         # print(resp)
#         objs = resp['total_rank']
#         qns = resp['questions']
#         submissions = resp['submissions']
#
#
#         #score array is updating here
#         # for index in range(0,4):
#         #     score_array[index] = qns[index]['credit']
#
#
#         if len(objs) == 0:
#             break
#
#         for index,obj in enumerate(objs):
#             # print(obj['username'], obj['rank'], calculate_solved_programs(obj['score']))
#             obj['username'] = obj['username'].lower()
#
#
#             if obj['username'] in users:
#                 curr_name = obj['username']
#
#                 print(curr_name)
#                 curr_obj_solved = f'{len(submissions[index])}/4'
#
#                 curr_obj = get_dictionary(obj, curr_obj_solved)
#
#
#                 final_list.append(curr_obj)  # no use have to be removed
#
#                 # updating attendees
#                 attended.append(curr_name)  # nno use instead of this have an integer for calculate no of attendees
#
#                 sheet_obj.cell(row=user_row[curr_name], column=col_to_fill_rank).fill = PatternFill(start_color=green,
#                                                                                                     end_color=green,
#                                                                                                     fill_type="solid")
#                 sheet_obj.cell(row=user_row[curr_name], column=col_to_fill_solved).fill = PatternFill(start_color=green,
#                                                                                                       end_color=green,
#                                                                                                       fill_type="solid")
#
#                 # update in sheet
#                 sheet_obj.cell(row=user_row[curr_name], column=col_to_fill_rank).value = curr_obj['rank']
#                 sheet_obj.cell(row=user_row[curr_name], column=col_to_fill_solved).value = curr_obj['solved']
#
#             # print(obj)
#         print('page = ', page, ' done')
#         page = page + 1
#     # break  # temporary purpose
#     # accessing final list
#     for participated_user_info in final_list:
#         print(participated_user_info)
#
#     # fill nA for not attended (already filled above so commented)
#     # for user in users:
#     #     if user not in attended:
#     #         sheet_obj.cell(row=user_row[user], column=col_to_fill_rank).value = 'NA'
#     #         sheet_obj.cell(row=user_row[user], column=col_to_fill_solved).value = 'NA'
#     #         print('filled at', user_row[user], col_to_fill_rank, col_to_fill_solved)
#
#     # saving as same file
#     wb_obj.save(path)
#     # if you want to save as different file use below method
#     # wb_obj.save("C:\\Users\\nobel\\PycharmProjects\\pythonProject\\III CSE A leetcode.xlsx")
#
# except Exception as e:
#     print(e)
#
# saving as same file
wb_obj.save(path)
print(time.time() - start_time)
