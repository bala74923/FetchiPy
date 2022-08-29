import openpyxl

wb = openpyxl.Workbook()
ws = wb.create_sheet('new sheet')
print(wb.sheetnames)

fake_sheet = wb['Sheet']
wb.remove(fake_sheet)
print(wb.sheetnames)

#wb.remove(wb['Sheet'])

wb.save('newfile.xlsx')