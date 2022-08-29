# from bs4 import BeautifulSoup  not used so removed
import os
import time

import requests
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.chart import PieChart, Reference,series
from openpyxl.chart.series import DataPoint

start_time = time.time()
CONTEST_NAME = "Weekly contest 307"
#contest_name = "weekly-contest-307"  # inpghggut
contest_name = CONTEST_NAME.lower().strip().replace(' ','-')
# score_array = [3,4,5,6]  # input will be updated soon
path = "C:\\Users\\nobel\\PycharmProjects\\pythonProject\\INPUT"  # path where input is fetched
output_path = "C:\\Users\\nobel\\PycharmProjects\\pythonProject\\OUTPUT"
dir_list = os.listdir(path)
path_list = []
for file in dir_list:
    path_list.append( path+'\\'+file)
for path in path_list:
    print(path)

rank_book = openpyxl.Workbook()
department_rank_book = openpyxl.Workbook()
chart_book  = openpyxl.Workbook()

# creating sheets


#path_list = ["C:\\Users\\nobel\\PycharmProjects\\pythonProject\\LC Student ID.xlsx","C:\\Users\\nobel\\PycharmProjects\\pythonProject\\III CSE A leetcode nd.xlsx"]

black = '00000000'
red = 'f54242'
green = '42f572'
yellow = 'ebcf34'
violet = '6d51a6'

thin_border = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000')
)

def fetch_file_name(path):
    return (path.split("\\")[-1]).split(".")[0]




def searchMaxCol(leet_code_id_row,startCol):
    for col_val in range(startCol,sheet_obj.max_column+1):
        curr_val = sheet_obj.cell(row=leet_code_id_row, column=col_val).value
        if curr_val is None:
            return col_val-1
    return sheet_obj.max_column

def get_dictionary(obj_from_list,solved_progs):
    mydict = {
        #"name": obj_from_list['username'],
        "rank": obj_from_list['rank'],
        "solved":solved_progs
    }
    return mydict

def is_contains_leetcode_names(column,row):
    cell = str(sheet_obj.cell(row=row, column=column).value).lower()
    return cell.__contains__("name")

def is_contains_leetcode_ids(string):
    string = str(string).lower()

    # return true if cell has string leetcode id
    return string.__contains__("user") and string.__contains__("id")

def get_row_col_position_for_leetcode_names(row_val):
    for col in range(1,sheet_obj.max_column+1):
        if is_contains_leetcode_names(col,row_val):
            return col

def get_row_col_position_for_leetcode_id():
    for row_ind in range(1, max_row + 1):
        for col_ind in range(1, sheet_obj.max_column + 1):
            curr_val = sheet_obj.cell(row=row_ind, column=col_ind).value
            if curr_val is not None and is_contains_leetcode_ids(curr_val):
                return [row_ind, col_ind]
    return None


def create_columns():
    # we have to create column names for ranking and solved
    # color
    sheet_obj.cell(row=leetcode_id_cell_position[0], column=col_to_fill_rank).fill = PatternFill(start_color=yellow,
                                                                                                 end_color=yellow,
                                                                                                 fill_type="solid")
    sheet_obj.cell(row=leetcode_id_cell_position[0], column=col_to_fill_solved).fill = PatternFill(start_color=yellow,
                                                                                                   end_color=yellow,
                                                                                                   fill_type="solid")
    # name
    sheet_obj.cell(row=leetcode_id_cell_position[0], column=col_to_fill_rank).value = 'Ranking'
    sheet_obj.cell(row=leetcode_id_cell_position[0], column=col_to_fill_solved).value = 'Solved'

    # border
    sheet_obj.cell(row=leetcode_id_cell_position[0], column=col_to_fill_rank).border = thin_border
    sheet_obj.cell(row=leetcode_id_cell_position[0], column=col_to_fill_solved).border = thin_border

def create_header_cell(class_sheet_obj,row,col,color,header_name):
    cell =  class_sheet_obj.cell(row=row, column=col)
    cell.fill = PatternFill(start_color=yellow,end_color=yellow,fill_type="solid")
    cell.value = header_name
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

def get_department_name(name):
    return name.strip().lower().split(" ")[0]

def make_class_sheet(class_sheet_obj,list,class_name):
    start_row = 3
    start_col = 5
    end_col = start_col+5  # starts from 0 so end at 5 equals total 6
    header_row = start_row+1
    values_start_row = header_row+1

    for col_val in range(start_col,end_col+1):
        class_sheet_obj.cell(row=start_row, column=col_val).border = thin_border

    class_sheet_obj.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    cell = class_sheet_obj.cell(row=start_row, column=start_col)
    cell.value = class_name
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # headers created
    header_names = ['SI', 'USERNAME', 'USER ID', 'SECTION', 'RANK', 'SOLVED']
    for index,header_name in enumerate(header_names):
        create_header_cell(class_sheet_obj, header_row, start_col+index,yellow,header_name)

    #values added
    for index,user_details  in enumerate(list):
        for detail_id in range(0,len(header_names)):
            cell = class_sheet_obj.cell(values_start_row+index,start_col+detail_id)

            cell.border = thin_border
            if index<=2: #top 3
                cell.fill = PatternFill(start_color=red, end_color=red, fill_type="solid")

            if detail_id==0: # SI
                cell.value = index+1
            else: # user name 0,user id 1,rank 2,solved 3
                cell.value = user_details[detail_id-1]
            cell.alignment = Alignment(horizontal='center', vertical='center')

def create_chart_for_department_details(year_chart_sheet,stats):
    for row in stats:
        year_chart_sheet.append(row)
    pie = PieChart()
    labels = Reference(year_chart_sheet, min_col=1, min_row=2, max_row=len(stats))
    data = Reference(year_chart_sheet, min_col=2, min_row=1, max_row=len(stats))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "ATTENDANCE PERCENTAGE "+CONTEST_NAME

    slice = DataPoint(idx=0, explosion=20)
    pie.series[0].data_points = [slice]

    year_chart_sheet.add_chart(pie, "D1")

def mention_contest_name(row,col):
    print('row = ', row, 'col = ', col)
    sheet_obj.cell(row=row, column=col).border = thin_border
    sheet_obj.cell(row=row, column=col+1).border = thin_border

    #sheet_obj.cell(row=row, column=col).value = contest_name

    sheet_obj.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    print(sheet_obj.merged_cells.ranges)

    cell = sheet_obj.cell(row=row, column=col)
    cell.value = contest_name


    cell.border = thin_border
    #
    cell.alignment = Alignment(horizontal='center', vertical='center')

#global
total_rank_dict = dict()
total_ranks = []
total_submissions = []

user_row = dict()  # maps users with his row, so we can later update it



final_list = []

# code to fetch total_ranks,total_submissions
page = 1
while True:
    try:
        API_URL_FMT = 'https://leetcode.com/contest/api/ranking/{}/?pagination={}&region=global'
        url = API_URL_FMT.format(contest_name, page)
        resp = requests.get(url).json()
        # print(resp)
        objs = resp['total_rank']
        qns = resp['questions']
        submissions = resp['submissions']
        if len(objs) == 0:
            break
        total_ranks.append(objs)
        total_submissions.append(qns)

        for index,obj in enumerate(objs):
            # print(obj['username'], obj['rank'], calculate_solved_programs(obj['score']))
            obj['username'] = obj['username'].lower()
            curr_obj_solved = f'{len(submissions[index])}/4'
            user_details = get_dictionary(obj, curr_obj_solved)
            total_rank_dict[obj['username']] = user_details
            # print(obj)
        print('page = ', page, ' done')
        page = page + 1
    except Exception as e:
        print(e,' so we cannot fetch details')
        break # some times program stop executing

for user in total_rank_dict.keys():
    print(user,total_rank_dict[user])


college_list =[]
college_rank_sheet_name = "College Toppers"
college_rank_sheet = rank_book.create_sheet(college_rank_sheet_name)
for path in path_list:
    wb_obj = openpyxl.load_workbook(path)
    #sheet_obj = wb_obj.active

    year_workbook_name = fetch_file_name(path)
    year_list = []
    year_rank_sheet_name = year_workbook_name+" Toppers"
    year_rank_sheet = rank_book.create_sheet(year_rank_sheet_name)

    department_dictionary = dict()
    for sheet_name in wb_obj.sheetnames:
        sheet_obj = wb_obj[sheet_name]
        print(sheet_name, sheet_obj)
        max_row = sheet_obj.max_row
        max_col = sheet_obj.max_column

        users = []  # for all username
        attended = []  # users who attended

        # finding all positions to fill (no user given input)
        leetcode_id_cell_position = get_row_col_position_for_leetcode_id()
        leetcode_id_column_name = leetcode_id_cell_position[1]  # 6-starts from 1
        leetcode_id_row_start_position = leetcode_id_cell_position[0] + 1  # 4- starts from 1
        leetcode_name_column = get_row_col_position_for_leetcode_names(leetcode_id_cell_position[0])
        max_col = searchMaxCol(leetcode_id_cell_position[0],leetcode_id_column_name)
        col_to_fill_rank = max_col + 1
        col_to_fill_solved = max_col + 2

        print(leetcode_id_cell_position)
        print(max_col)

        mention_contest_name(leetcode_id_cell_position[0] - 1, max_col + 1)
        create_columns()  # create columns

        class_list = []
        class_total = 0
        class_rank_sheet_name = year_workbook_name+' '+sheet_name+' Toppers'
        class_rank_sheet = rank_book.create_sheet(class_rank_sheet_name)
        for row_val in range(leetcode_id_row_start_position, max_row + 1):
            curr_name = str(sheet_obj.cell(row=row_val, column=leetcode_id_column_name).value)
            curr_name = curr_name.lower().strip()  # case not sensitive
            curr_user_name = str(sheet_obj.cell(row=row_val, column= leetcode_name_column).value)

            class_total+=1
            rank_cell = sheet_obj.cell(row=row_val, column=col_to_fill_rank)
            solved_cell = sheet_obj.cell(row=row_val, column=col_to_fill_solved)
            # default red for all
            rank_cell.fill = PatternFill(start_color=red, end_color=red, fill_type="solid")
            solved_cell.fill = PatternFill(start_color=red, end_color=red, fill_type="solid")

            # default NA fill for ALL
            rank_cell.value = 'Not Attended'
            solved_cell.value = 'Not Attended'

            # default black border for ALL
            rank_cell.border = thin_border
            solved_cell.border = thin_border

            # default center alignment for ALL
            rank_cell.alignment = Alignment(horizontal='center', vertical='center')
            solved_cell.alignment = Alignment(horizontal='center', vertical='center')

            # print(row_val, curr_name)
            if total_rank_dict.get(curr_name,None) is not None:
                #curr_name = curr_name.lower().strip()
                #users.append(curr_name)
                current_person = [curr_user_name, curr_name,
                                  year_workbook_name+' '+sheet_name,int(total_rank_dict[curr_name]['rank']) ,
                                  total_rank_dict[curr_name]['solved']]

                class_list.append(current_person)
                year_list.append(current_person)
                college_list.append(current_person)

                print(row_val, curr_name)
                # default red for all
                rank_cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
                solved_cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")

                # default NA fill for ALL
                rank_cell.value = total_rank_dict[curr_name]['rank']
                solved_cell.value = total_rank_dict[curr_name]['solved']
            # else:
            #     url = f'https://leetcode.com/graphql/?query=query{{ userContestRankingHistory(username: "{curr_name}") {{ attended trendDirection problemsSolved totalProblems finishTimeInSeconds rating ranking contest {{ title startTime }} }} }}'
            #     try:
            #         resp = requests.get(url).json()['errors'] # if errors is present then that is invalid
            #         # default red for all
            #         rank_cell.fill = PatternFill(start_color=violet, end_color=violet, fill_type="solid")
            #         solved_cell.fill = PatternFill(start_color=violet, end_color=violet, fill_type="solid")
            #
            #         # default NA fill for ALL
            #         rank_cell.value = 'Invalid ID'
            #         solved_cell.value = 'Invalid ID'
            #     except Exception as e:           #     if errors not present then valid
            #         pass

            #  print(users)
        class_list.sort(key=lambda x: x[3])
        print(' class wise list :')
        for index, curr_person_details in enumerate(class_list):
            print(index, curr_person_details)
        make_class_sheet(class_rank_sheet,class_list,class_rank_sheet_name+' '+CONTEST_NAME)

        # we have to add every class to particular department in dictionary
        department_name = get_department_name(sheet_name)
        # if previously department not readed then intialise empty list
        if department_name not in department_dictionary.keys():
            department_dictionary[department_name] = dict()
            department_dictionary[department_name]['list'] = []
            department_dictionary[department_name]['participated'] = 0
            department_dictionary[department_name]['total'] = 0
        # extend the class list with department

        department_dictionary[department_name]['list'].extend(class_list)
        department_dictionary[department_name]['participated']+= len(class_list)
        department_dictionary[department_name]['total'] += class_total

        #department_dictionary[department_name].extend(class_list)

    year_list.sort(key=lambda x: x[3])
    print(' year wise list :')
    for index, curr_person_details in enumerate(year_list):
        print(index, curr_person_details)
    make_class_sheet(year_rank_sheet, year_list, year_rank_sheet_name+' '+CONTEST_NAME)
    wb_obj.save(output_path+'\\'+year_workbook_name+'.xlsx') # should be saved because there are multiple workbook objects


    department_stats = [['department','participation']]
    year_chart_sheet = chart_book.create_sheet(year_workbook_name)

    for department_name in department_dictionary.keys():
        departement_sheet_name = year_workbook_name +' '+department_name
        department_dictionary[department_name]['list'].sort(key=lambda x: x[3])
        department_rank_sheet = department_rank_book.create_sheet(departement_sheet_name)
        make_class_sheet(department_rank_sheet,department_dictionary[department_name]['list'],
                         departement_sheet_name+' Toppers '+CONTEST_NAME)

        percentage = round(department_dictionary[department_name]['participated']*100/department_dictionary[department_name]['total'],2)
        department_stats.append([departement_sheet_name , percentage])

    create_chart_for_department_details(year_chart_sheet,department_stats)

college_list.sort(key=lambda x: x[3])
print('college wise list:')
for index,curr_person_details in enumerate(college_list):
    print(index,curr_person_details)
make_class_sheet(college_rank_sheet,college_list,college_rank_sheet_name+' '+CONTEST_NAME)

# remove empty sheet
department_rank_book.remove(department_rank_book['Sheet'])
department_rank_book.save(output_path+'\\Department RankBook.xlsx')

#remove empty sheet
rank_book.remove(rank_book['Sheet'])
rank_book.save(output_path+'\\RankBook.xlsx') #should be saved at end

#remove empty sheet
chart_book.remove(chart_book['Sheet'])
chart_book.save(output_path+'\\ChartBook.xlsx')

print(time.time() - start_time)
