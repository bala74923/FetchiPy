# from bs4 import BeautifulSoup  not used so removed
import math
import os
import time
# new add start
import _Chart
import _Leetcode
import _ContestName_asColumn
import colors
import last_page  # for get last page method
import threading
# import Threads
# new add end
import json
import requests
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.chart import PieChart, Reference, series
from openpyxl.chart.series import DataPoint

# added max row and column width adjustment than results with page link
start_time = time.time()
CONTEST_NAME = "biweekly contest 89"
# contest_name = "weekly-contest-307"  # inpghggut
contest_name = CONTEST_NAME.lower().strip().replace(' ', '-')
# score_array = [3,4,5,6]  # input will be updated soon
path = "D:\\MiniProject\\FetchiPy\\INPUT"  # path where input is fetched
output_path = "D:\\MiniProject\\FetchiPy\\OUTPUT\\"
output_list = [] #output names
last_page_of_contest = last_page.get_last_page_of_contest(contest_name)

dir_list = os.listdir(path)
path_list = []
for file in dir_list:
    path_list.append(path + '\\' + file)
for path in path_list:
    print(path)

rank_book = openpyxl.Workbook()
department_rank_book = openpyxl.Workbook()
chart_book = openpyxl.Workbook()

# creating sheets


# path_list = ["C:\\Users\\nobel\\PycharmProjects\\pythonProject\\LC Student ID.xlsx","C:\\Users\\nobel\\PycharmProjects\\pythonProject\\III CSE A leetcode nd.xlsx"]




def fetch_file_name(path):
    return (path.split("\\")[-1]).split(".")[0]


def get_dictionary(obj_from_list, solved_progs, page_link):
    mydict = {
        # "name": obj_from_list['username'],
        "rank": obj_from_list['rank'],
        "solved": solved_progs,
        "page_link": page_link
    }
    return mydict


def create_header_cell(class_sheet_obj, row, col, color, header_name):
    cell = class_sheet_obj.cell(row=row, column=col)
    cell.fill = PatternFill(start_color=colors.yellow, end_color=colors.yellow,
                            fill_type="solid")
    cell.value = header_name
    cell.border = colors.thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')


def create_columns(sheet_obj, leetcode_id_row_position, col_to_fill_rank, col_to_fill_solved, col_to_fill_page_link):
    # we have to create column names for ranking and solved
    create_header_cell(class_sheet_obj=sheet_obj, row=leetcode_id_row_position, col=col_to_fill_rank,
                       color=colors.yellow, header_name='Ranking')
    create_header_cell(class_sheet_obj=sheet_obj, row=leetcode_id_row_position, col=col_to_fill_solved,
                       color=colors.yellow, header_name='Solved')
    create_header_cell(class_sheet_obj=sheet_obj, row=leetcode_id_row_position, col=col_to_fill_page_link,
                       color=colors.yellow, header_name='Page Link')


def get_department_name(name):
    name = str(name).strip().lower()
    if name.__contains__("&"):  # AI & DS
        return name
    return name.split(" ")[0]


def make_class_sheet(class_sheet_obj, list, class_name):
    start_row = 3
    start_col = 5
    end_col = start_col + 6  # starts from 0 so end at 5 equals total 6
    header_row = start_row + 1
    values_start_row = header_row + 1
    end_row = header_row + len(list)

    for col_val in range(start_col, end_col + 1):
        class_sheet_obj.cell(row=start_row, column=col_val).border = colors.thin_border

    class_sheet_obj.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    cell = class_sheet_obj.cell(row=start_row, column=start_col)
    cell.value = class_name
    cell.border = colors.thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # headers created
    header_names = ['SI', 'USERNAME', 'USER ID', 'SECTION', 'RANK', 'SOLVED', 'PAGE LINK']
    for index, header_name in enumerate(header_names):
        create_header_cell(class_sheet_obj, header_row, start_col + index, colors.yellow, header_name)

    # values added
    for index, user_details in enumerate(list):
        for detail_id in range(0, len(header_names)):
            cell = class_sheet_obj.cell(values_start_row + index, start_col + detail_id)

            cell.border = colors.thin_border
            if index <= 2:  # top 3
                cell.fill = PatternFill(start_color=colors.red, end_color=colors.red, fill_type="solid")

            if detail_id == 0:  # SI
                cell.value = index + 1
            else:  # user name 0,user id 1,rank 2,solved 3
                cell.value = user_details[detail_id - 1]
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # once values are added should be adjusted
    _Chart.adjust_column_width(startRow=header_row, startCol=start_col, maxRow=end_row, maxCol=end_col, ws=class_sheet_obj)










# global
total_rank_dict = dict()
total_ranks = []
total_submissions = []

user_row = dict()  # maps users with his row, so we can later update it

final_list = []

# code to fetch total_ranks,total_submissions

# start1
def iterate_pages(start,end):
    page = start
    while page<=end:
        try:
            API_URL_FMT = 'https://leetcode.com/contest/api/ranking/{}/?pagination={}&region=global'
            page_link = f'https://leetcode.com/contest/{contest_name}/ranking/{page}/'

            url = API_URL_FMT.format(contest_name, page)
            print(url)
            resp = requests.get(url).json()
            # print(resp)
            objs = resp['total_rank']
            qns = resp['questions']
            submissions = resp['submissions']
            if len(objs) == 0:
                break
            total_ranks.append(objs)
            total_submissions.append(qns)

            for index,obj in enumerate(objs):
                # print(obj['username'], obj['rank'], calculate_solved_programs(obj['score']))
                obj['username'] = obj['username'].lower()
                curr_obj_solved = f'{len(submissions[index])}/4'
                user_details = get_dictionary(obj, curr_obj_solved,page_link)
                total_rank_dict[obj['username']] = user_details
                # print(obj)
            print('page = ', page, ' done')
            page = page + 1
        except Exception as e:
            print(e,' so we cannot fetch details',page,url)
            #break # some times program stop executing

    for user in total_rank_dict.keys():
        print(user,total_rank_dict[user])

# for example create some thread 3 and boost up
# lets create pairs
'''
def giveMeThreads(totalPages):
    interval =  1# every thread 10 pages
    threadNo = int(math.ceil(totalPages/interval))
    mylist = []
    start ,end = 1,interval
    for i in range(1,threadNo+1): # goes from 1 to threadNumber
        print(start,end)
        mylist.append( threading.Thread(target=iterate_pages,args=(start,end,)))
        start = end+1
        end+=interval
        end = min(end,totalPages) # for the last page
    # for last thread
   # mylist.append(threading.Thread(target=iterate_pages, args=(start,totalPages ,)))
    return mylist

def startThreads(list):
    for t in list:
        t.start()

def joinThreads(list):
    for t in list:
        t.join()   # end1
interval_pages = int(last_page_of_contest/9)
thred_work_list = [ [1,interval_pages],
                    [interval_pages+1,2*interval_pages]
                    ,[2*interval_pages+1,3*interval_pages]
                    ,[3*interval_pages+1,4*interval_pages]
                    ,[4*interval_pages+1,5*interval_pages]
                    ,[5*interval_pages+1,6*interval_pages]
                    ,[6*interval_pages+1,7*interval_pages]
                    ,[7*interval_pages+1,8*interval_pages]
                    ,[8*interval_pages+1,last_page_of_contest]
                    ]


t1 = threading.Thread(target=iterate_pages,args=(thred_work_list[0][0] , thred_work_list[0][1],))
t2 = threading.Thread(target=iterate_pages,args=(thred_work_list[1][0] , thred_work_list[1][1],))
t3 = threading.Thread(target=iterate_pages,args=(thred_work_list[2][0] , thred_work_list[2][1],))
t4 = threading.Thread(target=iterate_pages,args=(thred_work_list[3][0] , thred_work_list[3][1],))
t5 = threading.Thread(target=iterate_pages,args=(thred_work_list[4][0] , thred_work_list[4][1],))
t6 = threading.Thread(target=iterate_pages,args=(thred_work_list[5][0] , thred_work_list[5][1],))
t7 = threading.Thread(target=iterate_pages,args=(thred_work_list[6][0] , thred_work_list[6][1],))
t8 = threading.Thread(target=iterate_pages,args=(thred_work_list[7][0] , thred_work_list[7][1],))
t9 = threading.Thread(target=iterate_pages,args=(thred_work_list[8][0] , thred_work_list[8][1],))
# t10 = threading.Thread(target=iterate_pages,args=(thred_work_list[9][0] , thred_work_list[9][1],))

t1.start()
t2.start()
t3.start()
t4.start()
t5.start()
t6.start()
t7.start()
t8.start()
t9.start()
# t10.start()

# wait for the threads
t1.join()
t2.join()
t3.join()
t4.join()
t5.join()
t6.join()
t7.join()
t8.join()
t9.join()
# t10.join() '''

# start2
'''
thread_list  = giveMeThreads(last_page_of_contest)
startThreads(thread_list)
joinThreads(thread_list) '''


# end 2


# new try for 1000 pages 1000 threads ->100 seconds
def iterate_pages(start):
    page = start
    while page == start:
        try:
            API_URL_FMT = 'https://leetcode.com/contest/api/ranking/{}/?pagination={}&region=global'
            page_link = f'https://leetcode.com/contest/{contest_name}/ranking/{page}/'

            url = API_URL_FMT.format(contest_name, page)
            print(url)
            resp = requests.get(url).json()
            # print(resp)
            objs = resp['total_rank']
            qns = resp['questions']
            submissions = resp['submissions']
            # if len(objs) == 0: # as already we are parsing over valid pages
            #    break
            total_ranks.append(objs)
            total_submissions.append(qns)

            for index, obj in enumerate(objs):
                # print(obj['username'], obj['rank'], calculate_solved_programs(obj['score']))
                obj['username'] = obj['username'].lower()
                curr_obj_solved = f'{len(submissions[index])}/4'
                user_details = get_dictionary(obj, curr_obj_solved, page_link)
                total_rank_dict[obj['username']] = user_details
                # print(obj)
            print('page = ', page, ' done')
            page = page + 1
        except Exception as e:
            print(e, ' so we cannot fetch details', page, url)
            time.sleep(0.1)
            # break # some times program stop executing

    #for user in total_rank_dict.keys():
     #   print(user, total_rank_dict[user])


# for example create some thread 3 and boost up
# lets create pairs

def giveMeThreads(totalPages):
    mylist = []
    for page_no in range(1, totalPages + 1):  # goes from 1 to threadNumber
        print(page_no)
        mylist.append(threading.Thread(target=iterate_pages, args=(page_no,)))
    return mylist


def startThreads(list):
    for t in list:
        t.start()


def joinThreads(list):
    for t in list:
        t.join() 


thread_list = giveMeThreads(last_page_of_contest)
startThreads(thread_list)
joinThreads(thread_list)
# end of new try

college_list = []
college_rank_sheet_name = "College Toppers"
college_rank_sheet = rank_book.create_sheet(college_rank_sheet_name)
for path in path_list:
    wb_obj = openpyxl.load_workbook(path)
    # sheet_obj = wb_obj.active

    year_workbook_name = fetch_file_name(path)
    year_list = []
    year_rank_sheet_name = year_workbook_name + " Toppers"
    year_rank_sheet = rank_book.create_sheet(year_rank_sheet_name)

    department_dictionary = dict()
    for sheet_name in wb_obj.sheetnames:
        sheet_obj = wb_obj[sheet_name]
        print(sheet_name, sheet_obj)
        max_row = sheet_obj.max_row
        max_col = sheet_obj.max_column

        users = []  # for all username
        attended = []  # users who attended

        # finding all positions to fill (no user given input)
        leetcode_id_cell_position = _Leetcode.get_row_col_position_for_leetcode_id(sheet_obj=sheet_obj)
        leetcode_id_column_name = leetcode_id_cell_position[1]  # 6-starts from 1
        leetcode_id_row_start_position = leetcode_id_cell_position[0] + 1  # 4- starts from 1
        leetcode_name_column = _Leetcode.get_row_col_position_for_leetcode_names(leetcode_id_cell_position[0],sheet_obj=sheet_obj)
        max_col = _Leetcode.searchMaxCol(leetcode_id_cell_position[0], leetcode_id_column_name,sheet_obj)
        max_row = _Leetcode.searchMaxRow(startRow=leetcode_id_cell_position[0], leet_code_id_col=leetcode_id_column_name,sheet_obj=sheet_obj)
        col_to_fill_rank = max_col + 1
        col_to_fill_solved = max_col + 2
        col_to_fill_page_link = max_col + 3

        print(leetcode_id_cell_position)
        print(max_col)

        _ContestName_asColumn.mention_contest_name(leetcode_id_cell_position[0] - 1, max_col + 1,
                                                   sheet_obj=sheet_obj,contest_name=CONTEST_NAME)

        create_columns(sheet_obj=sheet_obj, leetcode_id_row_position=leetcode_id_cell_position[0],
                       col_to_fill_rank=col_to_fill_rank, col_to_fill_solved=col_to_fill_solved,
                       col_to_fill_page_link=col_to_fill_page_link)  # create columns

        class_list = []
        class_total = 0
        class_rank_sheet_name = year_workbook_name + ' ' + sheet_name + ' Toppers'
        class_rank_sheet = rank_book.create_sheet(class_rank_sheet_name)
        for row_val in range(leetcode_id_row_start_position, max_row + 1):
            curr_name = str(sheet_obj.cell(row=row_val, column=leetcode_id_column_name).value)
            curr_name = curr_name.lower().strip()  # case not sensitive
            curr_user_name = str(sheet_obj.cell(row=row_val, column=leetcode_name_column).value)

            class_total += 1
            rank_cell = sheet_obj.cell(row=row_val, column=col_to_fill_rank)
            solved_cell = sheet_obj.cell(row=row_val, column=col_to_fill_solved)
            page_link_cell = sheet_obj.cell(row=row_val, column=col_to_fill_page_link)
            # default red for all
            rank_cell.fill = PatternFill(start_color=colors.red, end_color=colors.red, fill_type="solid")
            solved_cell.fill = PatternFill(start_color=colors.red, end_color=colors.red, fill_type="solid")
            page_link_cell.fill = PatternFill(start_color=colors.red, end_color=colors.red, fill_type="solid")

            # default NA fill for ALL
            rank_cell.value = 'Not Attended'
            solved_cell.value = 'Not Attended'
            page_link_cell.value = 'Not Attended'

            # default black border for ALL
            rank_cell.border = colors.thin_border
            solved_cell.border = colors.thin_border
            page_link_cell.border = colors.thin_border

            # default center alignment for ALL
            rank_cell.alignment = Alignment(horizontal='center', vertical='center')
            solved_cell.alignment = Alignment(horizontal='center', vertical='center')
            page_link_cell.alignment = Alignment(horizontal='center', vertical='center')

            # print(row_val, curr_name)
            if total_rank_dict.get(curr_name, None) is not None:
                # curr_name = curr_name.lower().strip()
                # users.append(curr_name)
                current_person = [curr_user_name, curr_name,
                                  year_workbook_name + ' ' + sheet_name, int(total_rank_dict[curr_name]['rank']),
                                  total_rank_dict[curr_name]['solved'], total_rank_dict[curr_name]['page_link']]

                class_list.append(current_person)
                year_list.append(current_person)
                college_list.append(current_person)

                print(row_val, curr_name)
                # default red for all
                rank_cell.fill = PatternFill(start_color=colors.green, end_color=colors.green, fill_type="solid")
                solved_cell.fill = PatternFill(start_color=colors.green, end_color=colors.green, fill_type="solid")
                page_link_cell.fill = PatternFill(start_color=colors.green, end_color=colors.green, fill_type="solid")

                # default NA fill for ALL
                rank_cell.value = total_rank_dict[curr_name]['rank']
                solved_cell.value = total_rank_dict[curr_name]['solved']
                page_link_cell.value = total_rank_dict[curr_name]['page_link']
            '''else: //for invalid
                url = f'https://leetcode.com/graphql/?query=query{{ userContestRankingHistory(username: "{curr_name}") {{ attended trendDirection problemsSolved totalProblems finishTimeInSeconds rating ranking contest {{ title startTime }} }} }}'
                try:
                    resp = requests.get(url).json()['errors'] # if errors is present then that is invalid
                    # default red for all
                    rank_cell.fill = PatternFill(start_color=violet, end_color=violet, fill_type="solid")
                    solved_cell.fill = PatternFill(start_color=violet, end_color=violet, fill_type="solid")
            
                    # default NA fill for ALL
                    rank_cell.value = 'Invalid ID'
                    solved_cell.value = 'Invalid ID'
                except Exception as e:           #     if errors not present then valid
                    pass '''

            # print(users)
        _Chart.adjust_column_width(startRow=leetcode_id_cell_position[0], startCol=leetcode_id_cell_position[1],
                            maxRow=max_row, maxCol=col_to_fill_page_link, ws=sheet_obj)  # for normal class sheets
        class_list.sort(key=lambda x: x[3])
        print(' class wise list :')
        for index, curr_person_details in enumerate(class_list):
            print(index, curr_person_details)
        make_class_sheet(class_sheet_obj=class_rank_sheet, list=class_list,
                         class_name=class_rank_sheet_name + ' ' + CONTEST_NAME)

        # we have to add every class to particular department in dictionary
        department_name = get_department_name(sheet_name)
        # if previously department not readed then intialise empty list
        if department_name not in department_dictionary.keys():
            department_dictionary[department_name] = dict()
            department_dictionary[department_name]['list'] = []
            department_dictionary[department_name]['participated'] = 0
            department_dictionary[department_name]['total'] = 0
        # extend the class list with department

        department_dictionary[department_name]['list'].extend(class_list)
        department_dictionary[department_name]['participated'] += len(class_list)
        department_dictionary[department_name]['total'] += class_total

        # department_dictionary[department_name].extend(class_list)

    year_list.sort(key=lambda x: x[3])
    print(' year wise list :')
    for index, curr_person_details in enumerate(year_list):
        print(index, curr_person_details)
    make_class_sheet(year_rank_sheet, year_list, year_rank_sheet_name + ' ' + CONTEST_NAME)

    wb_file_name = output_path + '\\' + year_workbook_name + '.xlsx'
    wb_obj.save(wb_file_name)  # should be saved because there are multiple workbook objects
    output_list.append(wb_file_name)


    department_stats = [['department', 'participated', 'total', 'percentage']]
    year_chart_sheet = chart_book.create_sheet(year_workbook_name)

    for department_name in department_dictionary.keys():
        departement_sheet_name = year_workbook_name + ' ' + department_name
        department_dictionary[department_name]['list'].sort(key=lambda x: x[3])
        department_rank_sheet = department_rank_book.create_sheet(departement_sheet_name)

        make_class_sheet(department_rank_sheet, department_dictionary[department_name]['list'],
                         departement_sheet_name + ' Toppers ' + CONTEST_NAME)

        # refer https://stackoverflow.com/questions/42344041/how-to-copy-worksheet-from-one-workbook-to-another-one-using-openpyxl
        # making every department sheet as separate woorkbooks
        department_wise_workbook = openpyxl.Workbook()
        # department_wise_workbook_name = departement_sheet_name+' Toppers '+CONTEST_NAME
        department_wise_workbook_name = departement_sheet_name
        department_wise_sheet = department_wise_workbook.active
        make_class_sheet(department_wise_sheet, department_dictionary[department_name]['list'],
                         departement_sheet_name + ' Toppers ' + CONTEST_NAME)
        department_wise_workbook.save(output_path + '\\' + department_wise_workbook_name + '.xlsx')

        percentage = round(
            department_dictionary[department_name]['participated'] * 100 / department_dictionary[department_name][
                'total'], 2)
        department_stats.append([departement_sheet_name, department_dictionary[department_name]['participated'],
                                 department_dictionary[department_name]['total'], percentage, ])

    _Chart.create_chart_for_department_details(year_chart_sheet, department_stats,CONTEST_NAME)

college_list.sort(key=lambda x: x[3])
print('college wise list:')
for index, curr_person_details in enumerate(college_list):
    print(index, curr_person_details)
make_class_sheet(college_rank_sheet, college_list, college_rank_sheet_name + ' ' + CONTEST_NAME)

# remove empty sheet
department_rank_book.remove(department_rank_book['Sheet'])
department_file_name =output_path + '\\Department RankBook.xlsx'
department_rank_book.save(department_name)
output_list.append(department_file_name)

# remove empty sheet
rank_book.remove(rank_book['Sheet'])
rank_book_name = output_path + '\\RankBook.xlsx'
rank_book.save(rank_book_name)  # should be saved at end
output_list.append(rank_book_name)

# remove empty sheet
chart_book.remove(chart_book['Sheet'])
chart_book_name = output_path + '\\ChartBook.xlsx'
chart_book.save(chart_book_name)
output_list.append(chart_book_name)

for output in output_list:
    print(output)
print(time.time() - start_time)