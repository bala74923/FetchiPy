import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment

def mention_contest_name(row, col,sheet_obj,contest_name):
    print('row = ', row, 'col = ', col)
    sheet_obj.cell(row=row, column=col).border = colors.thin_border
    sheet_obj.cell(row=row, column=col + 1).border = colors.thin_border
    sheet_obj.cell(row=row, column=col + 2).border = colors.thin_border

    # sheet_obj.cell(row=row, column=col).value = contest_name

    sheet_obj.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
    print(sheet_obj.merged_cells.ranges)

    cell = sheet_obj.cell(row=row, column=col)
    cell.value = contest_name

    cell.border = colors.thin_border
    #
    cell.alignment = Alignment(horizontal='center', vertical='center')
