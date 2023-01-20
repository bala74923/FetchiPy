# from bs4 import BeautifulSoup  not used so removed
# create ranksheet for each class,year,college
import math
import os
import time
import json
import requests
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.styles.borders import BORDER_THIN
import static
import threading

start_time = time.time()
#CONTEST_NAME = "Weekly contest 307"
#contest_name = "weekly-contest-307"  # inpghggut
#contest_name = CONTEST_NAME.lower().strip().replace(' ','-')

# score_array = [3,4,5,6]  # input will be updated soon
path = "D:\\MiniProject\\FetchiPy\\INPUT"  # path where input is fetched
output_path = "D:\\MiniProject\\FetchiPy\\OUTPUT"
dir_list = os.listdir(path)
path_list = []
for file in dir_list:
    path_list.append( path+'\\'+file)
for path in path_list:
    print(path)

# creating sheets


#path_list = ["C:\\Users\\nobel\\PycharmProjects\\pythonProject\\LC Student ID.xlsx","C:\\Users\\nobel\\PycharmProjects\\pythonProject\\III CSE A leetcode nd.xlsx"]

black = '00000000'
red = 'f54242'
green = '42f572'
yellow = 'ebcf34'
violet = '6d51a6'

thin_border = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000')
)

def fetch_file_name(path):
    return (path.split("\\")[-1]).split(".")[0]

def adjust_column_width(startRow,startCol,maxRow,maxCol,ws):
    dims = {}
    for col_val in range(startCol, maxCol + 1):
        max_width = 8
        for row_val in range(startRow, maxRow + 1):
            cell = ws.cell(row=row_val, column=col_val)
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                print(cell.value)
                # max_width = max(len(str(cell.value)), max_width)
    for col, value in dims.items():
        ws.column_dimensions[col].width = value+4 # for rank only

def searchMaxCol(leet_code_id_row,startCol):
    for col_val in range(startCol,sheet_obj.max_column+1):
        curr_val = sheet_obj.cell(row=leet_code_id_row, column=col_val).value
        if curr_val is None:
            return col_val-1
    return sheet_obj.max_column

def searchMaxRow(startRow,leet_code_id_col):
    for row_val in range(startRow,sheet_obj.max_row+1):
        curr_val = sheet_obj.cell(row=row_val, column=leet_code_id_col).value
        if curr_val is None:
            return row_val-1
    return sheet_obj.max_row

def get_dictionary(obj_from_list,solved_progs):
    mydict = {
        #"name": obj_from_list['username'],
        "rank": obj_from_list['rank'],
        "solved":solved_progs
    }
    return mydict

def is_contains_leetcode_names(column,row):
    cell = str(sheet_obj.cell(row=row, column=column).value).lower()
    return cell.__contains__("name")

def is_contains_leetcode_ids(string):
    string = str(string).lower()

    # return true if cell has string leetcode id
    return string.__contains__("user") and string.__contains__("id")

def get_row_col_position_for_leetcode_names(row_val):
    for col in range(1,sheet_obj.max_column+1):
        if is_contains_leetcode_names(col,row_val):
            return col

def get_row_col_position_for_leetcode_id():
    for row_ind in range(1, max_row + 1):
        for col_ind in range(1, sheet_obj.max_column + 1):
            curr_val = sheet_obj.cell(row=row_ind, column=col_ind).value
            if curr_val is not None and is_contains_leetcode_ids(curr_val):
                return [row_ind, col_ind]
    return None

def create_columns(sheet_obj,leetcode_id_row_position,new_col_number, name_of_col):
    # we have to create column names for ranking and solved
    # color
    create_header_cell(class_sheet_obj=sheet_obj,row=leetcode_id_row_position,
                       col= new_col_number,color=yellow ,header_name= name_of_col)

def create_header_cell(class_sheet_obj,row,col,color,header_name):
    cell =  class_sheet_obj.cell(row=row, column=col)
    cell.fill = PatternFill(start_color=yellow,end_color=yellow,fill_type="solid")
    cell.value = header_name
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

def adjust_column_width(startRow,startCol,maxRow,maxCol,ws):
    dims = {}
    for col_val in range(startCol, maxCol + 1):
        max_width = 8
        for row_val in range(startRow, maxRow + 1):
            cell = ws.cell(row=row_val, column=col_val)
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                print(cell.value)
                # max_width = max(len(str(cell.value)), max_width)
    for col, value in dims.items():
        ws.column_dimensions[col].width = value+4 # for rank only

#global
total_rank_dict = dict()
total_ranks = []
total_submissions = []

user_row = dict()  # maps users with his row, so we can later update it

final_list = []
for path in path_list:
    wb_obj = openpyxl.load_workbook(path)
    wb_name = fetch_file_name(path)
    #sheet_obj = wb_obj.active
    for sheet_name in wb_obj.sheetnames:
        sheet_obj = wb_obj[sheet_name]
        print(sheet_name, sheet_obj)
        max_row = sheet_obj.max_row
        max_col = sheet_obj.max_column

        users = []  # for all username
        attended = []  # users who attended

        # finding all positions to fill (no user given input)
        leetcode_id_cell_position = get_row_col_position_for_leetcode_id()
        leetcode_id_column_name = leetcode_id_cell_position[1]  # 6-starts from 1
        leetcode_id_row_start_position = leetcode_id_cell_position[0] + 1  # 4- starts from 1
        leetcode_name_column = get_row_col_position_for_leetcode_names(leetcode_id_cell_position[0])
        max_col = searchMaxCol(leetcode_id_cell_position[0],leetcode_id_column_name)
        max_row = searchMaxRow(startRow=leetcode_id_cell_position[0],leet_code_id_col=leetcode_id_column_name)

        #col_to_fill_valid = max_col + 1

        col_to_fill_rating = max_col +1 ;
        col_to_fill_attended = col_to_fill_rating + 1;
        col_to_fill_global_rank  = col_to_fill_attended+1;
        col_to_fill_top_percentage = col_to_fill_global_rank +1 ;

        col_to_fill_all_solved = col_to_fill_top_percentage +1 ;
        col_to_fill_easy_solved = col_to_fill_all_solved+ 1;
        col_to_fill_medium_solved = col_to_fill_easy_solved +1 ;
        col_to_fill_hard_solved = col_to_fill_medium_solved + 1;

        print(leetcode_id_cell_position)
        print(max_col)

       # create_columns(sheet_obj=sheet_obj, leetcode_id_row_position=leetcode_id_cell_position[0],new_col_number=col_to_fill_valid, name_of_col='valid check')  # create columns
        create_columns(sheet_obj=sheet_obj, leetcode_id_row_position=leetcode_id_cell_position[0],
                       new_col_number=col_to_fill_rating, name_of_col='Contest Rating')
        create_columns(sheet_obj=sheet_obj, leetcode_id_row_position=leetcode_id_cell_position[0],
                       new_col_number=col_to_fill_attended, name_of_col='Contest Attended')
        create_columns(sheet_obj=sheet_obj, leetcode_id_row_position=leetcode_id_cell_position[0],
                       new_col_number=col_to_fill_global_rank, name_of_col='Global Rank')
        create_columns(sheet_obj=sheet_obj, leetcode_id_row_position=leetcode_id_cell_position[0],
                       new_col_number=col_to_fill_top_percentage, name_of_col='Top Percentage')

        create_columns(sheet_obj=sheet_obj, leetcode_id_row_position=leetcode_id_cell_position[0],
                       new_col_number=col_to_fill_all_solved, name_of_col='ALL')
        create_columns(sheet_obj=sheet_obj, leetcode_id_row_position=leetcode_id_cell_position[0],
                       new_col_number=col_to_fill_easy_solved, name_of_col='Easy')
        create_columns(sheet_obj=sheet_obj, leetcode_id_row_position=leetcode_id_cell_position[0],
                       new_col_number=col_to_fill_medium_solved, name_of_col='Medium')
        create_columns(sheet_obj=sheet_obj, leetcode_id_row_position=leetcode_id_cell_position[0],
                       new_col_number=col_to_fill_hard_solved, name_of_col='Hard')

        #  fill sheet is the process ->here
        # def process(items, start, end):
        #     for item in items[start:end]:
        #         try:
        #             api.my_operation(item)
        #         except Exception:
        #             print('error with item')




        row_numbers_to_be_processed = []
        for row_value_to_be_processed in range(leetcode_id_row_start_position, max_row+1):
            row_numbers_to_be_processed.append(row_value_to_be_processed)

        #while splitting itself fill_sheet function is triggered


        #[start,end) -> end is excluded always
        def fill_sheet(row_numbers_to_be_processed , row_start , row_end):
            for row_val in row_numbers_to_be_processed[row_start:row_end]:
                curr_name = str(sheet_obj.cell(row=row_val, column=leetcode_id_column_name).value)
                curr_name = curr_name.lower().strip()  # case not sensitive
                curr_user_name = str(sheet_obj.cell(row=row_val, column= leetcode_name_column).value)
                print(curr_user_name+' =>'+sheet_name)

                #try to fetch details about user
                try:
                    #valid_cell = sheet_obj.cell(row=row_val, column=col_to_fill_valid)
                    all_solved_cell  = sheet_obj.cell(row=row_val, column=col_to_fill_all_solved)
                    easy_solved_cell = sheet_obj.cell(row=row_val, column=col_to_fill_easy_solved)
                    medium_solved_cell = sheet_obj.cell(row=row_val, column=col_to_fill_medium_solved)
                    hard_solved_cell = sheet_obj.cell(row=row_val, column= col_to_fill_hard_solved)

                    contest_rating_cell = sheet_obj.cell(row=row_val, column= col_to_fill_rating)
                    contest_attended_cell = sheet_obj.cell(row=row_val, column= col_to_fill_attended)
                    global_rank_cell = sheet_obj.cell(row=row_val, column=col_to_fill_global_rank)
                    top_percentage_cell =sheet_obj.cell(row=row_val, column= col_to_fill_top_percentage)

                    new_cells_list = [contest_rating_cell, contest_attended_cell ,global_rank_cell , top_percentage_cell,
                                      all_solved_cell , easy_solved_cell , medium_solved_cell , hard_solved_cell]


                    # default red for all
                    #valid_cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
                    for new_cell in new_cells_list:
                        new_cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
                        new_cell.border = thin_border
                        new_cell.alignment = Alignment(horizontal='center', vertical='center')



                    # default valid fill for ALL -> now fill the current value
                    #valid_cell.value = 'VALID'

                    # default black border for ALL
                   # valid_cell.border = thin_border -> line 224
                    # default center alignment for ALL
                   # valid_cell.alignment = Alignment(horizontal='center', vertical='center') --> line 225

                    url = f'https://leetcode.com/graphql/?query=query{{ userContestRankingHistory(username: "{curr_name}") {{ attended trendDirection problemsSolved totalProblems finishTimeInSeconds rating ranking contest {{ title startTime }} }} }}'

                    query = static.solved_count_query

                    variables = {'username': curr_name}

                    url = 'https://leetcode.com/graphql/'
                    r = requests.post(url, json={'query': query, 'variables': variables})

                    json_data_count = json.loads(r.text)

                    usernameHandle = json_data_count['data']['matchedUser']['username']
                    total = json_data_count['data']['matchedUser']['submitStats']['acSubmissionNum'][0]['count']
                    easy = json_data_count['data']['matchedUser']['submitStats']['acSubmissionNum'][1]['count']
                    med = json_data_count['data']['matchedUser']['submitStats']['acSubmissionNum'][2]['count']
                    hard = json_data_count['data']['matchedUser']['submitStats']['acSubmissionNum'][3]['count']

                    # entering values to sheet
                    all_solved_cell.value = total
                    easy_solved_cell.value = easy
                    medium_solved_cell.value = med
                    hard_solved_cell.value = hard

                    # contest raing query
                    query = static.contest_info_query
                    request2 = requests.post(url, json={'query': query, 'variables': variables})
                    json_data_rating = json.loads( request2 .text)

                    contest_rating_of_user = json_data_rating['data']['userContestRanking']['rating'];
                    contest_rating_of_user = int(math.ceil(contest_rating_of_user))

                    global_rank_of_user =json_data_rating['data']['userContestRanking']['globalRanking'] ;
                    contest_attended_by_user = json_data_rating['data']['userContestRanking']['attendedContestsCount'] ;
                    top_percentage_of_user = json_data_rating['data']['userContestRanking']['topPercentage'] ;
                    total_participants  = json_data_rating['data']['userContestRanking']['totalParticipants'];

                    #entering values of rankings
                    contest_rating_cell.value = contest_rating_of_user
                    contest_attended_cell.value = contest_attended_by_user
                    global_rank_cell.value =global_rank_of_user
                    top_percentage_cell.value = top_percentage_of_user

                    #print('user = '+curr_user_name)
                    print(f'name :{curr_user_name} solved:{total} rating :{contest_rating_of_user}')

                #invalid user
                except Exception as e:
                    #resp = requests.get(url).json()['errors'] # if errors is present then that is invalid
                    # default red for all
                    #valid_cell.fill = PatternFill(start_color=violet, end_color=violet, fill_type="solid")
                    for new_cell in new_cells_list:
                        new_cell.value = 'Invalid ID'
                        new_cell.fill = PatternFill(start_color=violet, end_color=violet, fill_type="solid")
                    # default NA fill for ALL
                    #valid_cell.value = 'Invalid ID'

        def split_processing(items, num_splits=4):
            split_size = len(items) // num_splits
            threads = []
            for i in range(num_splits):
                # determine the indices of the list this thread will handle
                start = i * split_size
                # special case on the last chunk to account for uneven splits
                end = None if i + 1 == num_splits else (i + 1) * split_size
                # create the thread
                threads.append(
                    threading.Thread(target=fill_sheet, args=(items, start, end)))
                threads[-1].start()  # start the thread we just created

            # wait for all threads to finish
            for t in threads:
                t.join()


        split_processing(row_numbers_to_be_processed)

        adjust_column_width(startRow=leetcode_id_cell_position[0], startCol=leetcode_id_cell_position[1],
                            maxRow=max_row, maxCol=col_to_fill_rating, ws=sheet_obj)
        adjust_column_width(startRow=leetcode_id_cell_position[0], startCol=leetcode_id_cell_position[1],
                            maxRow=max_row, maxCol=col_to_fill_attended, ws=sheet_obj)
        adjust_column_width(startRow=leetcode_id_cell_position[0], startCol=leetcode_id_cell_position[1],
                            maxRow=max_row, maxCol=col_to_fill_global_rank, ws=sheet_obj)
        adjust_column_width(startRow=leetcode_id_cell_position[0], startCol=leetcode_id_cell_position[1],
                            maxRow=max_row, maxCol=col_to_fill_top_percentage, ws=sheet_obj)


        adjust_column_width(startRow=leetcode_id_cell_position[0], startCol=leetcode_id_cell_position[1],
                            maxRow=max_row, maxCol=col_to_fill_all_solved, ws=sheet_obj)
        adjust_column_width(startRow=leetcode_id_cell_position[0], startCol=leetcode_id_cell_position[1],
                            maxRow=max_row, maxCol=col_to_fill_easy_solved, ws=sheet_obj)
        adjust_column_width(startRow=leetcode_id_cell_position[0], startCol=leetcode_id_cell_position[1],
                            maxRow=max_row, maxCol=col_to_fill_medium_solved, ws=sheet_obj)
        adjust_column_width(startRow=leetcode_id_cell_position[0], startCol=leetcode_id_cell_position[1],
                            maxRow=max_row, maxCol=col_to_fill_hard_solved, ws=sheet_obj)

    wb_obj.save(output_path+'\\'+wb_name+'- profile details.xlsx') # should be saved because there are multiple workbook objects
print(time.time() - start_time)
