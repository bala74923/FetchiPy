# from bs4 import BeautifulSoup  not used so removed
# create ranksheet for each class,year,college
import os
import time

import requests
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.styles.borders import BORDER_THIN

start_time = time.time()
#CONTEST_NAME = "Weekly contest 307"
#contest_name = "weekly-contest-307"  # inpghggut
#contest_name = CONTEST_NAME.lower().strip().replace(' ','-')

# score_array = [3,4,5,6]  # input will be updated soon
path = "C:\\Users\\nobel\\PycharmProjects\\pythonProject\\INPUT"  # path where input is fetched
output_path = "C:\\Users\\nobel\\PycharmProjects\\pythonProject\\CONSOLE"
dir_list = os.listdir(path)
path_list = []
for file in dir_list:
    path_list.append( path+'\\'+file)
for path in path_list:
    print(path)

# creating sheets


#path_list = ["C:\\Users\\nobel\\PycharmProjects\\pythonProject\\LC Student ID.xlsx","C:\\Users\\nobel\\PycharmProjects\\pythonProject\\III CSE A leetcode nd.xlsx"]

black = '00000000'
red = 'f54242'
green = '42f572'
yellow = 'ebcf34'
violet = '6d51a6'

thin_border = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000')
)

def fetch_file_name(path):
    return (path.split("\\")[-1]).split(".")[0]

def adjust_column_width(startRow,startCol,maxRow,maxCol,ws):
    dims = {}
    for col_val in range(startCol, maxCol + 1):
        max_width = 8
        for row_val in range(startRow, maxRow + 1):
            cell = ws.cell(row=row_val, column=col_val)
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                print(cell.value)
                # max_width = max(len(str(cell.value)), max_width)
    for col, value in dims.items():
        ws.column_dimensions[col].width = value+4 # for rank only

def searchMaxCol(leet_code_id_row,startCol):
    for col_val in range(startCol,sheet_obj.max_column+1):
        curr_val = sheet_obj.cell(row=leet_code_id_row, column=col_val).value
        if curr_val is None:
            return col_val-1
    return sheet_obj.max_column

def searchMaxRow(startRow,leet_code_id_col):
    for row_val in range(startRow,sheet_obj.max_row+1):
        curr_val = sheet_obj.cell(row=row_val, column=leet_code_id_col).value
        if curr_val is None:
            return row_val-1
    return sheet_obj.max_row

def get_dictionary(obj_from_list,solved_progs):
    mydict = {
        #"name": obj_from_list['username'],
        "rank": obj_from_list['rank'],
        "solved":solved_progs
    }
    return mydict

def is_contains_leetcode_names(column,row):
    cell = str(sheet_obj.cell(row=row, column=column).value).lower()
    return cell.__contains__("name")

def is_contains_leetcode_ids(string):
    string = str(string).lower()

    # return true if cell has string leetcode id
    return string.__contains__("user") and string.__contains__("id")

def get_row_col_position_for_leetcode_names(row_val):
    for col in range(1,sheet_obj.max_column+1):
        if is_contains_leetcode_names(col,row_val):
            return col

def get_row_col_position_for_leetcode_id():
    for row_ind in range(1, max_row + 1):
        for col_ind in range(1, sheet_obj.max_column + 1):
            curr_val = sheet_obj.cell(row=row_ind, column=col_ind).value
            if curr_val is not None and is_contains_leetcode_ids(curr_val):
                return [row_ind, col_ind]
    return None

def create_columns(sheet_obj,leetcode_id_row_position,col_to_fill_valid):
    # we have to create column names for ranking and solved
    # color
    create_header_cell(class_sheet_obj=sheet_obj,row=leetcode_id_row_position,
                       col=col_to_fill_valid,color=yellow ,header_name= 'VALID CHECK')

def create_header_cell(class_sheet_obj,row,col,color,header_name):
    cell =  class_sheet_obj.cell(row=row, column=col)
    cell.fill = PatternFill(start_color=yellow,end_color=yellow,fill_type="solid")
    cell.value = header_name
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

def adjust_column_width(startRow,startCol,maxRow,maxCol,ws):
    dims = {}
    for col_val in range(startCol, maxCol + 1):
        max_width = 8
        for row_val in range(startRow, maxRow + 1):
            cell = ws.cell(row=row_val, column=col_val)
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                print(cell.value)
                # max_width = max(len(str(cell.value)), max_width)
    for col, value in dims.items():
        ws.column_dimensions[col].width = value+4 # for rank only

#global
total_rank_dict = dict()
total_ranks = []
total_submissions = []

user_row = dict()  # maps users with his row, so we can later update it

final_list = []
for path in path_list:
    wb_obj = openpyxl.load_workbook(path)
    wb_name = fetch_file_name(path)
    #sheet_obj = wb_obj.active
    for sheet_name in wb_obj.sheetnames:
        sheet_obj = wb_obj[sheet_name]
        print(sheet_name, sheet_obj)
        max_row = sheet_obj.max_row
        max_col = sheet_obj.max_column

        users = []  # for all username
        attended = []  # users who attended

        # finding all positions to fill (no user given input)
        leetcode_id_cell_position = get_row_col_position_for_leetcode_id()
        leetcode_id_column_name = leetcode_id_cell_position[1]  # 6-starts from 1
        leetcode_id_row_start_position = leetcode_id_cell_position[0] + 1  # 4- starts from 1
        leetcode_name_column = get_row_col_position_for_leetcode_names(leetcode_id_cell_position[0])
        max_col = searchMaxCol(leetcode_id_cell_position[0],leetcode_id_column_name)
        max_row = searchMaxRow(startRow=leetcode_id_cell_position[0],leet_code_id_col=leetcode_id_column_name)
        col_to_fill_valid = max_col + 1

        print(leetcode_id_cell_position)
        print(max_col)

        create_columns(sheet_obj=sheet_obj,leetcode_id_row_position=leetcode_id_cell_position[0]
                       ,col_to_fill_valid=col_to_fill_valid)  # create columns
        for row_val in range(leetcode_id_row_start_position, max_row + 1):
            curr_name = str(sheet_obj.cell(row=row_val, column=leetcode_id_column_name).value)
            curr_name = curr_name.lower().strip()  # case not sensitive
            curr_user_name = str(sheet_obj.cell(row=row_val, column= leetcode_name_column).value)
            print(curr_user_name+' =>'+sheet_name)
            valid_cell = sheet_obj.cell(row=row_val, column=col_to_fill_valid)
            # default red for all
            valid_cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
            # default valid fill for ALL
            valid_cell.value = 'VALID'
            # default black border for ALL
            valid_cell.border = thin_border
            # default center alignment for ALL
            valid_cell.alignment = Alignment(horizontal='center', vertical='center')

            url = f'https://leetcode.com/graphql/?query=query{{ userContestRankingHistory(username: "{curr_name}") {{ attended trendDirection problemsSolved totalProblems finishTimeInSeconds rating ranking contest {{ title startTime }} }} }}'
            try:
                resp = requests.get(url).json()['errors'] # if errors is present then that is invalid
                # default red for all
                valid_cell.fill = PatternFill(start_color=violet, end_color=violet, fill_type="solid")

                # default NA fill for ALL
                valid_cell.value = 'Invalid ID'
            except Exception as e:           #     if errors not present then valid
                pass
        adjust_column_width(startRow=leetcode_id_cell_position[0], startCol=leetcode_id_cell_position[1],
                            maxRow=max_row, maxCol=col_to_fill_valid, ws=sheet_obj)
    wb_obj.save(output_path+'\\'+wb_name+'.xlsx') # should be saved because there are multiple workbook objects
print(time.time() - start_time)
