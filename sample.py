# import requests
# from selenium import webdriver
# username = "user4029ok"
# # url = f'https://leetcode.com/graphql/?query=query{{ userContestRankingHistory(username: "{username}") {{ attended trendDirection problemsSolved totalProblems finishTimeInSeconds rating ranking contest {{ title startTime }} }} }}'
# #resp = requests.get(url).json()['errors']
# #print(resp)
# contest_name = 'weekly-contest-307'
# page = 1
# url = f'https://leetcode.com/contest/{contest_name}/ranking/{page}/'
#
# driver = webdriver.Firefox()
# driver.get(url)
# screenshot = driver.save_screenshot('my_screenshot.png')
# driver.quit()
# val = 13.45555
# val  = round(val,2)
# print(val)
from openpyxl.utils import get_column_letter
import openpyxl
'''
path = 'OUTPUT\\III year.xlsx'
output_path = 'OUTPUT\\III year space.xlsx'
wb = openpyxl.load_workbook(path)

def adjust_column_width(startRow,startCol,maxRow,maxCol,ws):
    dims = {}
    for col_val in range(startCol, maxCol + 1):
        max_width = 8
        for row_val in range(startRow, maxRow + 1):
            cell = ws.cell(row=row_val, column=col_val)
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                print(cell.value)
                # max_width = max(len(str(cell.value)), max_width)
    for col, value in dims.items():
        ws.column_dimensions[col].width = value


for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    adjust_column_width(4,9,64,11,ws)

wb.save(output_path)
'''
import requests
username = "user4029ok"
url = 'leetcode.com/graphql/?query=query { matchedUser(username: "user4029ok") { username submitStats: submitStatsGlobal { acSubmissionNum { difficulty count submissions } } } }'
resp = requests.get(url)
print(resp)
