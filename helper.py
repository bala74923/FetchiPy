import time

import requests
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.styles.borders import BORDER_THIN

start_time = time.time()
contest_name = "Weekly Contest 306"
contest_name.lower()
path = "C:\\Users\\nobel\\PycharmProjects\\pythonProject\\III CSE A leetcode.xlsx"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
max_row = sheet_obj.max_row
max_col = sheet_obj.max_column


black = '00000000'
red = 'f54242'
green = '42f572'
yellow = 'ebcf34'
blue = '3a96e0'
violet = '9e3ae0'

thin_border = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000')
)


def create_columns():
    # we have to create column names for ranking and solved
    # color
    sheet_obj.cell(row=leetcode_id_cell_position[0], column=col_to_fill_rank).fill = PatternFill(start_color=yellow,
                                                                                                 end_color=yellow,
                                                                                                 fill_type="solid")
    sheet_obj.cell(row=leetcode_id_cell_position[0], column=col_to_fill_solved).fill = PatternFill(start_color=yellow,
                                                                                                   end_color=yellow,
                                                                                                   fill_type="solid")
    # name
    sheet_obj.cell(row=leetcode_id_cell_position[0], column=col_to_fill_rank).value = 'Ranking'
    sheet_obj.cell(row=leetcode_id_cell_position[0], column=col_to_fill_solved).value = 'Solved'

    # border
    sheet_obj.cell(row=leetcode_id_cell_position[0], column=col_to_fill_rank).border = thin_border
    sheet_obj.cell(row=leetcode_id_cell_position[0], column=col_to_fill_solved).border = thin_border


def is_contains_leetcode_ids(string):
    string = string.lower()

    # return true if cell has string leetcode id
    return string.__contains__("leetcode") and string.__contains__("id")


def get_row_col_position_for_leetcode_id():
    for row_ind in range(1, max_row + 1):
        for col_ind in range(1, max_col + 1):
            curr_val = sheet_obj.cell(row=row_ind, column=col_ind).value
            if curr_val is not None and is_contains_leetcode_ids(curr_val):
                return [row_ind, col_ind]
    return None

# finding all positions to fill (no user given input)
leetcode_id_cell_position = get_row_col_position_for_leetcode_id()
leetcode_id_column_name = leetcode_id_cell_position[1]  # 6-starts from 1
leetcode_id_row_start_position = leetcode_id_cell_position[0] + 1  # 4- starts from 1
col_to_fill_rank = max_col + 1
col_to_fill_solved = max_col + 2

create_columns()
for row_val in range(leetcode_id_row_start_position, max_row + 1):
    username = sheet_obj.cell(row=row_val, column=leetcode_id_column_name).value
    ranking_cell = sheet_obj.cell(row=row_val, column=col_to_fill_rank)
    solved_cell = sheet_obj.cell(row=row_val, column=col_to_fill_solved)

    ranking_cell.border = thin_border
    solved_cell.border = thin_border

    if username is not None:
        try:
            url = f'https://leetcode.com/graphql/?query=query{{ userContestRankingHistory(username: "{username}") {{ attended trendDirection problemsSolved totalProblems finishTimeInSeconds rating ranking contest {{ title startTime }} }} }}'
            #print(url)
            resp = requests.get(url).json()
            hist = resp['data']['userContestRankingHistory']
            user_details = None
            print(hist[-1])
            for val in hist[::-1]:
                val['contest']['title'].lower()
                if contest_name == val['contest']['title']:
                    user_details = val
                    break

            if user_details['attended']:
                ranking_cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
                solved_cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")

                solved_cell.value = user_details['problemsSolved']
                ranking_cell.value = user_details['ranking']
            else:
                # default red for all
                ranking_cell.fill = PatternFill(start_color=red, end_color=red, fill_type="solid")
                solved_cell.fill = PatternFill(start_color=red, end_color=red, fill_type="solid")

                # default NA fill for ALL
                ranking_cell.value = 'cant detect'
                solved_cell.value = '0/4'

        except Exception as e: #violet
            ranking_cell.fill = PatternFill(start_color=violet, end_color=violet, fill_type="solid")
            solved_cell.fill = PatternFill(start_color=violet, end_color=violet, fill_type="solid")

            ranking_cell.value = 'Invalid id'
            solved_cell.value = 'Invalid id'
            print(username+' has given wrong id')

    # have to color it in different color
    else: #blue
        ranking_cell.fill = PatternFill(start_color=blue, end_color=blue, fill_type="solid")
        solved_cell.fill = PatternFill(start_color=blue, end_color=blue, fill_type="solid")

        ranking_cell.value = 'empty ID'
        solved_cell.value = 'empty ID'


wb_obj.save(path)

